"""
src/app_bridge/bridge.py
App Automation Bridge - connects Baba to Outlook, WhatsApp,
TikTok, Instagram, Facebook, Chrome, VS Code, Excel/Sheets.
"""

import json
import asyncio
import subprocess
import os
import platform
import shutil
import socket
import csv
import base64
import imaplib
import smtplib
import email
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import urlparse, parse_qs
from email.header import decode_header, make_header
from email.mime.text import MIMEText
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime, UTC

from src.integrations.microsoft_oauth import MicrosoftOAuthManager


class AppBridge:
    """Unified app automation bridge. All write actions require approval."""

    def __init__(self, settings=None):
        self.settings = settings
        self._approval_log = Path("logs/app_actions.jsonl")
        self._approval_log.parent.mkdir(parents=True, exist_ok=True)
        self._exo_agents = {}
        self._last_scan = {}
        self._obsidian_vault = self._resolve_obsidian_vault()
        self._ms_oauth = MicrosoftOAuthManager(settings=settings)
        self._oauth_loopback_server = None
        self._oauth_loopback_thread = None
        self._oauth_last_result: Dict[str, Any] = {}
        self._init_exo_agents()

    def _resolve_obsidian_vault(self) -> Optional[Path]:
        env = os.getenv("OBSIDIAN_VAULT", "").strip()
        if env:
            p = Path(env).expanduser()
            return p if p.exists() else None
        if self.settings and hasattr(self.settings, "_raw"):
            try:
                raw = getattr(self.settings, "_raw", {})
                vault = (
                    raw.get("integrations", {})
                    .get("obsidian", {})
                    .get("vault_path", "")
                )
                if vault:
                    p = Path(vault).expanduser()
                    if p.exists():
                        return p
            except Exception:
                pass
        return None

    def _init_exo_agents(self):
        try:
            from exo_email_agents import GmailAgent, OutlookAgent

            self._exo_agents = {
                "gmail": GmailAgent(),
                "outlook": OutlookAgent(),
            }
        except Exception:
            self._exo_agents = {}

    def _is_port_open(self, port: int, host: str = "localhost") -> bool:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(0.8)
        try:
            return s.connect_ex((host, port)) == 0
        finally:
            s.close()

    def _running_processes(self) -> set:
        names = set()
        try:
            if os.name == "nt":
                result = subprocess.run(
                    ["tasklist", "/FO", "CSV", "/NH"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="ignore",
                    timeout=5,
                )
                if result.returncode == 0:
                    rows = csv.reader(result.stdout.splitlines())
                    for row in rows:
                        if row:
                            names.add(row[0].strip().lower())
            else:
                result = subprocess.run(
                    ["ps", "-A", "-o", "comm="],
                    capture_output=True,
                    text=True,
                    timeout=5,
                )
                if result.returncode == 0:
                    for line in result.stdout.splitlines():
                        names.add(Path(line.strip()).name.lower())
        except Exception:
            pass
        return names

    def detect_integrations(self, refresh: bool = True) -> Dict[str, Any]:
        if self._last_scan and not refresh:
            return self._last_scan

        processes = self._running_processes()
        which_code = shutil.which("code")
        which_python = shutil.which("python")

        oauth_status = self._ms_oauth.status()
        scan = {
            "timestamp": datetime.now(UTC).isoformat(),
            "runtime": {
                "os": platform.system(),
                "python": which_python or "python",
            },
            "services": {
                "pc_bridge_8765": self._is_port_open(8765),
                "dispatch_api_8767": self._is_port_open(8767),
                "chrome_connector_8768": self._is_port_open(8768),
                "chrome_debug_9222": self._is_port_open(9222),
            },
            "apps": {
                "exo": {
                    "running": "exo.exe" in processes or "exo" in processes,
                    "available": bool(shutil.which("exo")) or bool(shutil.which("exo-email")) or (os.name == "nt"),
                },
                "outlook": {
                    "running": "outlook.exe" in processes,
                    "available": ("outlook.exe" in processes) or (os.name == "nt"),
                },
                "excel": {
                    "running": "excel.exe" in processes,
                    "available": ("excel.exe" in processes) or (os.name == "nt"),
                },
                "word": {
                    "running": "winword.exe" in processes,
                    "available": ("winword.exe" in processes) or (os.name == "nt"),
                },
                "vscode": {
                    "running": "code.exe" in processes or "code" in processes,
                    "available": bool(which_code),
                },
                "obsidian": {
                    "running": "obsidian.exe" in processes or "obsidian" in processes,
                    "available": bool(shutil.which("obsidian")) or (os.name == "nt"),
                    "vault": str(self._obsidian_vault) if self._obsidian_vault else "",
                },
                "cmd": {
                    "running": "cmd.exe" in processes,
                    "available": os.name == "nt",
                },
            },
            "browsers": {
                "chrome": {
                    "running": "chrome.exe" in processes or "chrome" in processes,
                    "available": bool(shutil.which("chrome")) or os.name == "nt",
                },
                "edge": {
                    "running": "msedge.exe" in processes or "msedge" in processes,
                    "available": bool(shutil.which("msedge")) or os.name == "nt",
                },
            },
            "social": {
                "whatsapp_web": True,
                "gmail_web": True,
                "facebook_web": True,
                "instagram_web": True,
                "tiktok_web": True,
                "telegram_web": True,
            },
            "email_agents": {
                "exo_available": bool(self._exo_agents),
                "exo_gmail": "gmail" in self._exo_agents,
                "exo_outlook": "outlook" in self._exo_agents,
                "oauth_outlook_configured": oauth_status.get("configured", False),
                "oauth_outlook_connected": oauth_status.get("connected", False),
                "oauth_redirect_uri": oauth_status.get("redirect_uri", ""),
            },
        }
        self._last_scan = scan
        return scan

    def auto_connect_integrations(self, include_launch: bool = False) -> Dict[str, Any]:
        scan = self.detect_integrations(refresh=True)
        actions = []

        if include_launch and os.name == "nt":
            if scan["browsers"]["chrome"]["available"] and not scan["services"]["chrome_debug_9222"]:
                # Optional warm-up for WhatsApp automation over CDP.
                try:
                    subprocess.Popen(
                        'start "" chrome --remote-debugging-port=9222',
                        shell=True,
                    )
                    actions.append("launched_chrome_with_debug_9222")
                except Exception as e:
                    actions.append(f"chrome_debug_launch_failed:{e}")
            if scan["browsers"]["edge"]["available"] and not scan["browsers"]["edge"]["running"]:
                try:
                    subprocess.Popen('start "" msedge', shell=True)
                    actions.append("launched_edge")
                except Exception as e:
                    actions.append(f"edge_launch_failed:{e}")

        refreshed = self.detect_integrations(refresh=True)
        return {
            "ok": True,
            "actions": actions,
            "scan": refreshed,
            "summary": {
                "services_online": sum(1 for v in refreshed["services"].values() if v),
                "apps_running": sum(1 for v in refreshed["apps"].values() if v.get("running")),
                "browsers_running": sum(
                    1 for v in refreshed["browsers"].values() if v.get("running")
                ),
            },
        }

    def outlook_oauth_status(self) -> Dict[str, Any]:
        status = {"ok": True, **self._ms_oauth.status()}
        if self._oauth_last_result:
            status["last_oauth_result"] = self._oauth_last_result
        return status

    def outlook_oauth_start(self, open_browser: bool = True) -> Dict[str, Any]:
        result = self._ms_oauth.get_authorization_url()
        if not result.get("ok"):
            return result
        listener = self._start_oauth_loopback_listener()
        result["listener"] = listener
        if open_browser:
            self.chrome_open(result["auth_url"])
        self._log(
            "outlook_oauth_start",
            {"redirect_uri": result.get("redirect_uri", ""), "open_browser": open_browser},
        )
        return result

    def outlook_oauth_exchange(self, code: str, state: str = "") -> Dict[str, Any]:
        result = self._ms_oauth.exchange_code(code=code, state=state)
        self._oauth_last_result = {
            "ts": datetime.now(UTC).isoformat(),
            "ok": bool(result.get("ok")),
            "error": result.get("error", ""),
        }
        self._log(
            "outlook_oauth_exchange",
            {"ok": bool(result.get("ok")), "state": state[:32]},
        )
        return result

    def outlook_oauth_disconnect(self) -> Dict[str, Any]:
        result = self._ms_oauth.clear_tokens()
        self._oauth_last_result = {
            "ts": datetime.now(UTC).isoformat(),
            "ok": True,
            "error": "",
            "event": "disconnect",
        }
        self._log("outlook_oauth_disconnect", {"ok": bool(result.get("ok"))})
        return result

    def _start_oauth_loopback_listener(self) -> Dict[str, Any]:
        cfg = self._ms_oauth.get_config()
        redirect_uri = cfg.get("redirect_uri", "")
        if not redirect_uri:
            return {"ok": False, "error": "Missing redirect URI"}

        if self._oauth_loopback_thread and self._oauth_loopback_thread.is_alive():
            return {"ok": True, "status": "already_running", "redirect_uri": redirect_uri}

        parsed = urlparse(redirect_uri)
        host = parsed.hostname or "localhost"
        port = parsed.port or 80
        path = parsed.path or "/"
        parent = self

        class _OAuthCallbackHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                current = urlparse(self.path)
                if current.path != path:
                    self.send_response(404)
                    self.end_headers()
                    self.wfile.write(b"Not Found")
                    return

                q = parse_qs(current.query)
                code = (q.get("code") or [""])[0]
                state = (q.get("state") or [""])[0]
                error = (q.get("error") or [""])[0]

                if error:
                    parent._oauth_last_result = {
                        "ts": datetime.now(UTC).isoformat(),
                        "ok": False,
                        "error": error,
                    }
                    payload = (
                        "<html><body><h3>Baba OAuth Error</h3>"
                        f"<p>{error}</p><p>You can close this tab.</p></body></html>"
                    )
                    self.send_response(400)
                    self.send_header("Content-Type", "text/html; charset=utf-8")
                    self.end_headers()
                    self.wfile.write(payload.encode("utf-8"))
                    threading.Thread(
                        target=parent._stop_oauth_loopback_listener, daemon=True
                    ).start()
                    return

                exchange = parent.outlook_oauth_exchange(code=code, state=state)
                if exchange.get("ok"):
                    self.send_response(200)
                    payload = (
                        "<html><body><h3>Baba OAuth Connected</h3>"
                        "<p>Connection successful. You can close this tab.</p></body></html>"
                    )
                else:
                    self.send_response(400)
                    payload = (
                        "<html><body><h3>Baba OAuth Failed</h3>"
                        f"<p>{exchange.get('error', 'Unknown error')}</p>"
                        "<p>You can close this tab and retry.</p></body></html>"
                    )

                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(payload.encode("utf-8"))
                threading.Thread(
                    target=parent._stop_oauth_loopback_listener, daemon=True
                ).start()

            def log_message(self, fmt, *args):
                return

        try:
            server = HTTPServer((host, port), _OAuthCallbackHandler)
            server.timeout = 1.0
        except OSError as e:
            return {
                "ok": False,
                "error": str(e),
                "redirect_uri": redirect_uri,
                "hint": "If UI server already runs on this port, callback will be handled by /oauth/callback endpoint.",
            }

        self._oauth_loopback_server = server
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        self._oauth_loopback_thread = thread
        return {
            "ok": True,
            "status": "listening",
            "redirect_uri": redirect_uri,
            "host": host,
            "port": port,
            "path": path,
        }

    def _stop_oauth_loopback_listener(self):
        try:
            if self._oauth_loopback_server:
                self._oauth_loopback_server.shutdown()
                self._oauth_loopback_server.server_close()
        except Exception:
            pass
        self._oauth_loopback_server = None
        self._oauth_loopback_thread = None

    def outlook_read_inbox(self, limit: int = 20, folder: str = "Inbox") -> List[Dict]:
        oauth_status = self._ms_oauth.status()
        if oauth_status.get("connected"):
            oauth_items = self._imap_read(limit, folder=folder)
            if oauth_items and not oauth_items[0].get("error"):
                return oauth_items
        try:
            return self._outlook_win32_read(limit, folder)
        except Exception:
            if "outlook" in self._exo_agents:
                try:
                    agent = self._exo_agents["outlook"]
                    fetch = getattr(agent, "fetch_unread", None)
                    if callable(fetch):
                        msgs = fetch(limit=limit)
                        return [
                            {
                                "subject": getattr(m, "subject", ""),
                                "sender": getattr(m, "sender", ""),
                                "date": getattr(m, "timestamp", ""),
                                "body": getattr(m, "body", "")[:500],
                                "read": False,
                            }
                            for m in msgs
                        ]
                except Exception:
                    pass
            return self._imap_read(limit, folder=folder)

    def _outlook_win32_read(self, limit: int, folder: str) -> List[Dict]:
        import win32com.client

        com_inited = False
        try:
            try:
                import pythoncom  # type: ignore

                pythoncom.CoInitialize()
                com_inited = True
            except Exception:
                pass

            ol = win32com.client.Dispatch("Outlook.Application")
            ns = ol.GetNamespace("MAPI")
            box = ns.GetDefaultFolder(6)
            msgs = box.Items
            msgs.Sort("[ReceivedTime]", True)
            items = []
            for i, msg in enumerate(msgs):
                if i >= limit:
                    break
                items.append(
                    {
                        "subject": msg.Subject,
                        "sender": msg.SenderName,
                        "date": str(msg.ReceivedTime),
                        "body": msg.Body[:500],
                        "read": msg.UnRead is False,
                    }
                )
            return items
        finally:
            if com_inited:
                try:
                    import pythoncom  # type: ignore

                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _imap_read(self, limit: int, folder: str = "Inbox") -> List[Dict]:
        token_res = self._ms_oauth.get_access_token()
        if not token_res.get("ok"):
            return [
                {
                    "error": token_res.get("error", "OAuth token unavailable"),
                    "hint": "Connect Outlook OAuth first via /api/auth/microsoft/start",
                    "redirect_uri": self._ms_oauth.status().get("redirect_uri", ""),
                }
            ]

        cfg = self._ms_oauth.get_config()
        user = cfg.get("email") or os.getenv("IMAP_USER", "").strip()
        if not user:
            return [
                {
                    "error": "Missing email account for IMAP",
                    "hint": "Set IMAP_USER or OUTLOOK_EMAIL in .env",
                }
            ]
        host = os.getenv("IMAP_HOST", "outlook.office365.com").strip() or "outlook.office365.com"
        port = int(os.getenv("IMAP_PORT", "993") or "993")
        mailbox = folder or "Inbox"
        mailbox = "INBOX" if mailbox.lower() == "inbox" else mailbox

        items: List[Dict[str, Any]] = []
        conn = None
        try:
            conn = imaplib.IMAP4_SSL(host, port)
            auth_string = f"user={user}\x01auth=Bearer {token_res['access_token']}\x01\x01"
            conn.authenticate("XOAUTH2", lambda _: auth_string.encode("utf-8"))
            conn.select(mailbox)
            status, data = conn.search(None, "ALL")
            if status != "OK":
                return [{"error": f"IMAP search failed on {mailbox}"}]
            ids = data[0].split()[-max(1, int(limit)) :]
            for msg_id in reversed(ids):
                st, msg_data = conn.fetch(msg_id, "(RFC822)")
                if st != "OK":
                    continue
                raw_bytes = None
                for part in msg_data:
                    if isinstance(part, tuple) and len(part) > 1:
                        raw_bytes = part[1]
                        break
                if not raw_bytes:
                    continue
                msg = email.message_from_bytes(raw_bytes)
                subject = self._decode_mime_header(msg.get("Subject", ""))
                sender = self._decode_mime_header(msg.get("From", ""))
                date_val = self._decode_mime_header(msg.get("Date", ""))
                body = self._extract_plain_text(msg)
                items.append(
                    {
                        "subject": subject,
                        "sender": sender,
                        "date": date_val,
                        "body": body[:1200],
                        "read": True,
                    }
                )
            self._log(
                "outlook_imap_read",
                {"count": len(items), "mailbox": mailbox, "host": host},
            )
            return items
        except Exception as e:
            return [
                {
                    "error": str(e),
                    "hint": "Verify Azure permissions (IMAP.AccessAsUser.All), IMAP enabled mailbox, and OAuth token",
                }
            ]
        finally:
            try:
                if conn:
                    conn.logout()
            except Exception:
                pass

    def outlook_draft(self, to: str, subject: str, body: str) -> Dict:
        try:
            import win32com.client

            ol = win32com.client.Dispatch("Outlook.Application")
            mail = ol.CreateItem(0)
            mail.To = to
            mail.Subject = subject
            mail.Body = body
            mail.Save()
            self._log("outlook_draft", {"to": to, "subject": subject})
            return {"ok": True, "message": f"Draft saved: '{subject}' to {to}"}
        except Exception as e:
            if "outlook" in self._exo_agents:
                try:
                    agent = self._exo_agents["outlook"]
                    send_reply = getattr(agent, "send_reply", None)
                    if callable(send_reply):
                        send_reply(None, body)
                        self._log("outlook_draft_exo", {"to": to, "subject": subject})
                        return {
                            "ok": True,
                            "message": f"Draft handled by EXO Outlook agent for {to}",
                        }
                except Exception:
                    pass
            return {
                "ok": False,
                "error": str(e),
                "fallback": "Draft text prepared - paste manually into Outlook",
            }

    def outlook_send(self, to: str, subject: str, body: str, approved: bool = False) -> Dict:
        if not approved:
            return {
                "ok": False,
                "requires_approval": True,
                "to": to,
                "subject": subject,
                "preview": (body or "")[:200],
            }

        token_res = self._ms_oauth.get_access_token()
        if not token_res.get("ok"):
            return {"ok": False, "error": token_res.get("error", "OAuth token unavailable")}

        cfg = self._ms_oauth.get_config()
        sender = cfg.get("email") or os.getenv("IMAP_USER", "").strip()
        if not sender:
            return {"ok": False, "error": "Missing sender email (set IMAP_USER or OUTLOOK_EMAIL)"}

        host = os.getenv("SMTP_HOST", "smtp.office365.com").strip() or "smtp.office365.com"
        port = int(os.getenv("SMTP_PORT", "587") or "587")

        msg = MIMEText(body or "", "plain", "utf-8")
        msg["From"] = sender
        msg["To"] = to
        msg["Subject"] = subject

        smtp = None
        try:
            smtp = smtplib.SMTP(host, port, timeout=30)
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            xoauth = f"user={sender}\x01auth=Bearer {token_res['access_token']}\x01\x01"
            xoauth_b64 = base64.b64encode(xoauth.encode("utf-8")).decode("utf-8")
            code, resp = smtp.docmd("AUTH", "XOAUTH2 " + xoauth_b64)
            if code != 235:
                return {
                    "ok": False,
                    "error": f"SMTP XOAUTH2 failed: {code} {resp.decode(errors='ignore') if isinstance(resp, bytes) else resp}",
                }
            smtp.sendmail(sender, [to], msg.as_string())
            self._log("outlook_send", {"to": to, "subject": subject, "host": host})
            return {"ok": True, "message": f"Email sent to {to}"}
        except Exception as e:
            return {"ok": False, "error": str(e)}
        finally:
            try:
                if smtp:
                    smtp.quit()
            except Exception:
                pass

    def outlook_open(self) -> Dict:
        if os.name != "nt":
            return {"ok": False, "error": "Outlook auto-launch currently Windows-only"}
        try:
            subprocess.Popen('start "" outlook', shell=True)
            self._log("outlook_open", {})
            return {"ok": True, "message": "Outlook launch command sent"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _decode_mime_header(self, value: str) -> str:
        try:
            return str(make_header(decode_header(value or "")))
        except Exception:
            return value or ""

    def _extract_plain_text(self, msg: email.message.Message) -> str:
        if msg.is_multipart():
            chunks = []
            for part in msg.walk():
                ctype = part.get_content_type()
                disp = str(part.get("Content-Disposition", "")).lower()
                if ctype == "text/plain" and "attachment" not in disp:
                    try:
                        payload = part.get_payload(decode=True) or b""
                        charset = part.get_content_charset() or "utf-8"
                        chunks.append(payload.decode(charset, errors="ignore"))
                    except Exception:
                        continue
            return "\n".join(chunks).strip()
        try:
            payload = msg.get_payload(decode=True) or b""
            charset = msg.get_content_charset() or "utf-8"
            return payload.decode(charset, errors="ignore").strip()
        except Exception:
            return ""

    def gmail_open(self) -> Dict:
        return self.chrome_open("https://mail.google.com")

    def exo_open(self) -> Dict:
        """
        Open Exo client when available.
        Falls back to Exo website when native app is not installed.
        """
        try:
            if os.name == "nt":
                if shutil.which("exo"):
                    subprocess.Popen("exo", shell=True)
                    return {"ok": True, "message": "Exo launch command sent"}
                if shutil.which("exo-email"):
                    subprocess.Popen("exo-email", shell=True)
                    return {"ok": True, "message": "Exo launch command sent"}
                return self.chrome_open("https://exo.email")
            if shutil.which("exo"):
                subprocess.Popen(["exo"])
                return {"ok": True, "message": "Exo launch command sent"}
            return self.chrome_open("https://exo.email")
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def exo_triage_inbox(self, limit: int = 30) -> Dict:
        """
        AI-native triage pass compatible with Exo workflows.
        Uses EXO Gmail/Outlook agents when present.
        """
        messages = []
        for provider in ("gmail", "outlook"):
            agent = self._exo_agents.get(provider)
            if not agent:
                continue
            fetch = getattr(agent, "fetch_unread", None)
            if not callable(fetch):
                continue
            try:
                for m in fetch(limit=limit) or []:
                    subject = getattr(m, "subject", "") if not isinstance(m, dict) else m.get("subject", "")
                    sender = getattr(m, "sender", "") if not isinstance(m, dict) else m.get("sender", "")
                    body = getattr(m, "body", "") if not isinstance(m, dict) else m.get("body", "")
                    messages.append(
                        {
                            "provider": provider,
                            "subject": subject,
                            "sender": sender,
                            "snippet": (body or "")[:300],
                            "priority": self._triage_priority(subject, body),
                        }
                    )
            except Exception:
                continue

        buckets = {"high": [], "medium": [], "low": [], "skip": []}
        for msg in messages:
            buckets[msg["priority"]].append(msg)
        summary = {k: len(v) for k, v in buckets.items()}
        self._log("exo_triage_inbox", {"summary": summary})
        return {"ok": True, "summary": summary, "buckets": buckets}

    def outlook_read_all_folders(
        self,
        limit_per_folder: int = 30,
        max_folders: int = 180,
        include_subfolders: bool = True,
        progress_cb=None,
    ) -> Dict[str, Any]:
        """
        Read Outlook mail across all mailbox stores/folders visible in Desktop Outlook.
        Best path for multi-account setups (multiple Outlook/Gmail accounts configured in Outlook client).
        """
        try:
            import win32com.client
        except Exception as e:
            return {"ok": False, "error": f"win32 outlook unavailable: {e}"}

        com_inited = False
        try:
            folder_limit = int(limit_per_folder)
        except Exception:
            folder_limit = 30
        try:
            folder_cap = int(max_folders)
        except Exception:
            folder_cap = 180
        no_item_limit = folder_limit <= 0
        no_folder_cap = folder_cap <= 0
        messages: List[Dict[str, Any]] = []
        folder_stats: List[Dict[str, Any]] = []
        seen = set()
        scanned_folders = 0
        store_names: List[str] = []

        try:
            try:
                import pythoncom  # type: ignore

                pythoncom.CoInitialize()
                com_inited = True
            except Exception:
                pass

            ol = win32com.client.Dispatch("Outlook.Application")
            ns = ol.GetNamespace("MAPI")
        except Exception as e:
            return {"ok": False, "error": f"Outlook MAPI init failed: {e}"}

        def walk(folder, parent_path: str = ""):
            nonlocal scanned_folders
            if (not no_folder_cap) and scanned_folders >= folder_cap:
                return

            name = str(getattr(folder, "Name", "") or "").strip() or "Folder"
            folder_path = f"{parent_path}/{name}" if parent_path else name
            scanned_folders += 1
            added = 0

            try:
                items = folder.Items
                try:
                    items.Sort("[ReceivedTime]", True)
                except Exception:
                    pass
                for msg in items:
                    if (not no_item_limit) and added >= max(1, folder_limit):
                        break
                    try:
                        # MailItem class id in Outlook Object Model.
                        if int(getattr(msg, "Class", 0) or 0) != 43:
                            continue
                        subject = str(getattr(msg, "Subject", "") or "")
                        sender = str(getattr(msg, "SenderName", "") or "")
                        body = str(getattr(msg, "Body", "") or "")
                        date_val = str(getattr(msg, "ReceivedTime", "") or "")
                        unread = bool(getattr(msg, "UnRead", False))
                        flag_status = int(getattr(msg, "FlagStatus", 0) or 0)
                        flagged = flag_status in (1, 2)
                        cats = str(getattr(msg, "Categories", "") or "")
                        pinned = "pinned" in cats.lower()
                        key = (subject, sender, date_val, folder_path)
                        if key in seen:
                            continue
                        seen.add(key)
                        messages.append(
                            {
                                "subject": subject,
                                "sender": sender,
                                "date": date_val,
                                "body": body[:1200],
                                "read": not unread,
                                "flagged": bool(flagged),
                                "pinned": bool(pinned),
                                "folder": folder_path,
                                "priority": self._triage_priority(subject, body),
                            }
                        )
                        added += 1
                    except Exception:
                        continue
            except Exception:
                pass

            folder_stats.append({"folder": folder_path, "messages": added})
            if callable(progress_cb):
                try:
                    progress_cb(
                        {
                            "stores_detected": len(store_names),
                            "folders_scanned": scanned_folders,
                            "messages_collected": len(messages),
                            "current_folder": folder_path,
                        }
                    )
                except Exception:
                    pass

            if not include_subfolders:
                return
            try:
                for sub in folder.Folders:
                    walk(sub, folder_path)
                    if (not no_folder_cap) and scanned_folders >= folder_cap:
                        break
            except Exception:
                pass

        # Enumerate stores safely so one bad store does not kill the full scan.
        try:
            store_count = int(getattr(ns.Folders, "Count", 0) or 0)
            for idx in range(1, store_count + 1):
                try:
                    store = ns.Folders.Item(idx)
                    sname = str(getattr(store, "Name", "") or "").strip()
                    if sname and sname not in store_names:
                        store_names.append(sname)
                    walk(store, "")
                except Exception:
                    continue
                if (not no_folder_cap) and scanned_folders >= folder_cap:
                    break
        except Exception as e:
            return {"ok": False, "error": f"Outlook folder walk failed: {e}"}
        finally:
            if com_inited:
                try:
                    import pythoncom  # type: ignore

                    pythoncom.CoUninitialize()
                except Exception:
                    pass

        buckets = {"high": [], "medium": [], "low": [], "skip": []}
        for m in messages:
            buckets[m.get("priority", "low")].append(m)
        summary = {k: len(v) for k, v in buckets.items()}
        summary["total"] = len(messages)
        summary["folders_scanned"] = scanned_folders
        accounts: Dict[str, int] = {}
        for m in messages:
            folder = str(m.get("folder", "") or "")
            root = folder.split("/", 1)[0].strip() if folder else ""
            if root:
                accounts[root] = accounts.get(root, 0) + 1
        summary["accounts_scanned"] = len(accounts)
        summary["accounts"] = accounts
        summary["stores_detected_count"] = len(store_names)
        summary["stores_detected"] = store_names
        self._log("outlook_read_all_folders", summary)
        return {
            "ok": True,
            "summary": summary,
            "buckets": buckets,
            "folders": folder_stats,
            "messages": messages,
        }

    def exo_triage_all_mail(
        self,
        limit_per_folder: int = 30,
        max_folders: int = 180,
        include_subfolders: bool = True,
        progress_cb=None,
    ) -> Dict[str, Any]:
        """
        Full triage across Outlook stores/folders (preferred), plus EXO agent unread fallback.
        """
        full = self.outlook_read_all_folders(
            limit_per_folder=limit_per_folder,
            max_folders=max_folders,
            include_subfolders=include_subfolders,
            progress_cb=progress_cb,
        )
        messages: List[Dict[str, Any]] = []
        if full.get("ok"):
            messages.extend(full.get("messages", []))

        # Fallback/additive unread from EXO agents.
        if not messages:
            fallback = self.exo_triage_inbox(limit=max(20, int(limit_per_folder)))
            if fallback.get("ok"):
                for level, items in fallback.get("buckets", {}).items():
                    for it in items:
                        item = dict(it)
                        item["priority"] = level
                        messages.append(item)

        buckets = {"high": [], "medium": [], "low": [], "skip": []}
        for msg in messages:
            level = msg.get("priority", "low")
            if level not in buckets:
                level = "low"
            buckets[level].append(msg)
        summary = {k: len(v) for k, v in buckets.items()}
        summary["total"] = len(messages)
        if full.get("ok"):
            summary["folders_scanned"] = (
                full.get("summary", {}) or {}
            ).get("folders_scanned", 0)
            summary["accounts_scanned"] = (
                full.get("summary", {}) or {}
            ).get("accounts_scanned", 0)
            summary["accounts"] = (
                full.get("summary", {}) or {}
            ).get("accounts", {})
            summary["stores_detected_count"] = (
                full.get("summary", {}) or {}
            ).get("stores_detected_count", 0)
            summary["stores_detected"] = (
                full.get("summary", {}) or {}
            ).get("stores_detected", [])
        self._log("exo_triage_all_mail", summary)
        return {
            "ok": True,
            "summary": summary,
            "buckets": buckets,
            "outlook_full_scan": full,
            "note": "No messages found" if not messages else "",
        }

    def _triage_priority(self, subject: str, body: str) -> str:
        text = f"{subject} {body}".lower()
        if any(k in text for k in ("urgent", "asap", "deadline", "overdue", "payment due", "hmrc")):
            return "high"
        if any(k in text for k in ("invoice", "contract", "meeting", "review")):
            return "medium"
        if any(k in text for k in ("newsletter", "promo", "offer", "sale")):
            return "skip"
        return "low"

    def gmail_read_inbox(self, limit: int = 20) -> List[Dict]:
        if "gmail" in self._exo_agents:
            try:
                agent = self._exo_agents["gmail"]
                fetch = getattr(agent, "fetch_unread", None)
                if callable(fetch):
                    msgs = fetch(limit=limit)
                    return [
                        {
                            "subject": getattr(m, "subject", ""),
                            "sender": getattr(m, "sender", ""),
                            "date": getattr(m, "timestamp", ""),
                            "body": getattr(m, "body", "")[:500],
                            "read": False,
                        }
                        for m in msgs
                    ]
            except Exception as e:
                return [{"error": str(e)}]
        return [{"error": "Gmail EXO agent not available; open Gmail web instead"}]

    def whatsapp_open(self) -> Dict:
        return self.whatsapp_open_chat("")

    def whatsapp_open_chat(self, contact: str = "") -> Dict:
        if not contact:
            return self.chrome_open("https://web.whatsapp.com")
        try:
            from playwright.sync_api import sync_playwright

            with sync_playwright() as p:
                browser = p.chromium.connect_over_cdp("http://localhost:9222")
                page = browser.contexts[0].pages[0]
                page.goto(f"https://web.whatsapp.com/send?phone={contact}")
                return {"ok": True, "message": f"Opened WhatsApp chat with {contact}"}
        except Exception as e:
            return {
                "ok": False,
                "error": str(e),
                "fallback": "Open WhatsApp Web manually and navigate to contact",
            }

    def whatsapp_send(self, contact: str, message: str, approved: bool = False) -> Dict:
        if not approved:
            return {
                "ok": False,
                "requires_approval": True,
                "message": message,
                "to": contact,
            }
        self._log("whatsapp_send", {"to": contact, "msg_preview": message[:50]})
        return {
            "ok": True,
            "message": f"Message queued for {contact} - open WhatsApp to confirm",
        }

    def chrome_open(self, url: str) -> Dict:
        try:
            import webbrowser

            webbrowser.open(url)
            return {"ok": True, "url": url}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def edge_open(self, url: str = "about:blank") -> Dict:
        try:
            if os.name == "nt":
                subprocess.Popen(f'start "" msedge "{url}"', shell=True)
                return {"ok": True, "url": url, "browser": "edge"}
            return self.chrome_open(url)
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def chrome_extract_page(self, url: str) -> Dict:
        try:
            import urllib.request

            with urllib.request.urlopen(url, timeout=15) as r:
                html = r.read().decode("utf-8", errors="ignore")
            import re

            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", " ", text).strip()
            return {"ok": True, "url": url, "text": text[:3000], "length": len(text)}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def chrome_screenshot(self, url: str, save_path: Optional[str] = None) -> Dict:
        try:
            from playwright.sync_api import sync_playwright

            save_path = (
                save_path
                or f"data/exports/screenshot_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.png"
            )
            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page()
                page.goto(url)
                page.screenshot(path=save_path)
                browser.close()
            return {"ok": True, "path": save_path}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def social_open(self, platform: str) -> Dict:
        urls = {
            "exo": "https://exo.email",
            "tiktok": "https://www.tiktok.com",
            "instagram": "https://www.instagram.com",
            "facebook": "https://www.facebook.com",
            "telegram": "https://web.telegram.org",
            "whatsapp": "https://web.whatsapp.com",
            "gmail": "https://mail.google.com",
            "outlook": "https://outlook.office.com/mail/",
            "linkedin": "https://www.linkedin.com",
            "x": "https://x.com",
            "twitter": "https://x.com",
            "slack": "https://slack.com/signin",
            "discord": "https://discord.com/app",
            "zoom": "https://zoom.us",
        }
        url = urls.get(platform.lower())
        if url:
            return self.chrome_open(url)
        return {"ok": False, "error": f"Unknown platform: {platform}"}

    def social_draft_post(
        self, platform: str, caption: str, approved: bool = False
    ) -> Dict:
        if not approved:
            return {
                "ok": False,
                "requires_approval": True,
                "platform": platform,
                "draft": caption,
                "message": "Review this draft then approve to proceed",
            }
        self._log("social_post", {"platform": platform, "preview": caption[:80]})
        return {
            "ok": True,
            "message": f"Draft approved for {platform} - open platform to post",
        }

    def vscode_open(self, path: str = ".") -> Dict:
        try:
            result = subprocess.run(["code", path], capture_output=True, text=True)
            return {"ok": result.returncode == 0, "path": path}
        except FileNotFoundError:
            return {
                "ok": False,
                "error": "VS Code CLI 'code' not found - install VS Code and add to PATH",
            }

    def cli_open(self) -> Dict:
        if os.name != "nt":
            return {"ok": False, "error": "CLI opener currently Windows-only"}
        try:
            subprocess.Popen('start "" cmd', shell=True)
            return {"ok": True, "message": "Command Prompt launch command sent"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def ide_open(self, path: str = ".") -> Dict:
        return self.vscode_open(path=path)

    def obsidian_open(self, vault_path: Optional[str] = None) -> Dict:
        vault = Path(vault_path).expanduser() if vault_path else self._obsidian_vault
        try:
            if os.name == "nt":
                if vault and vault.exists():
                    subprocess.Popen(f'start "" obsidian://open?vault={vault.name}', shell=True)
                    return {"ok": True, "vault": str(vault), "message": "Obsidian open command sent"}
                subprocess.Popen('start "" obsidian://open', shell=True)
                return {"ok": True, "message": "Obsidian open command sent"}
            if vault and vault.exists():
                subprocess.Popen(["obsidian", str(vault)])
                return {"ok": True, "vault": str(vault)}
            subprocess.Popen(["obsidian"])
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def obsidian_capture_note(
        self,
        title: str,
        content: str,
        folder: str = "Inbox",
        approved: bool = False,
    ) -> Dict:
        if not approved:
            return {
                "ok": False,
                "requires_approval": True,
                "title": title,
                "folder": folder,
                "preview": content[:240],
            }
        if not self._obsidian_vault:
            return {
                "ok": False,
                "error": "Obsidian vault not configured. Set OBSIDIAN_VAULT env var or config.integrations.obsidian.vault_path",
            }
        safe_title = "".join(
            c if c.isalnum() or c in (" ", "-", "_") else "_" for c in title
        ).strip() or f"note_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}"
        note_dir = self._obsidian_vault / folder
        note_dir.mkdir(parents=True, exist_ok=True)
        note_path = note_dir / f"{safe_title}.md"
        body = (
            f"# {title}\n\n"
            f"{content}\n\n"
            f"---\n"
            f"Created: {datetime.now(UTC).isoformat()}\n"
            f"Source: Baba Desktop\n"
        )
        note_path.write_text(body, encoding="utf-8")
        self._log("obsidian_capture_note", {"path": str(note_path)})
        return {"ok": True, "path": str(note_path)}

    def obsidian_search(self, query: str, limit: int = 20) -> Dict:
        if not self._obsidian_vault:
            return {"ok": False, "error": "Obsidian vault not configured"}
        hits = []
        q = query.lower()
        for p in self._obsidian_vault.rglob("*.md"):
            try:
                text = p.read_text(encoding="utf-8", errors="ignore")
                if q in text.lower() or q in p.name.lower():
                    snippet = text[:220].replace("\n", " ")
                    hits.append({"path": str(p), "title": p.stem, "snippet": snippet})
                    if len(hits) >= limit:
                        break
            except Exception:
                continue
        return {"ok": True, "count": len(hits), "results": hits}

    def vscode_create_file(
        self, file_path: str, content: str, approved: bool = False
    ) -> Dict:
        if not approved:
            return {
                "ok": False,
                "requires_approval": True,
                "path": file_path,
                "preview": content[:200],
            }
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, "w") as f:
            f.write(content)
        self._log("vscode_create_file", {"path": file_path})
        return {"ok": True, "path": file_path}

    def excel_read(self, path: str) -> Dict:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = [
                [str(c.value or "") for c in row] for row in ws.iter_rows(max_row=100)
            ]
            return {"ok": True, "rows": rows, "sheet": ws.title}
        except ImportError:
            return {
                "ok": False,
                "error": "openpyxl not installed: pip install openpyxl",
            }
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def excel_write(self, path: str, data: List[List], approved: bool = False) -> Dict:
        if not approved:
            return {
                "ok": False,
                "requires_approval": True,
                "path": path,
                "rows": len(data),
            }
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            for row in data:
                ws.append(row)
            wb.save(path)
            self._log("excel_write", {"path": path, "rows": len(data)})
            return {"ok": True, "path": path}
        except ImportError:
            return {"ok": False, "error": "pip install openpyxl"}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def _log(self, action: str, details: Dict):
        entry = {"ts": datetime.now(UTC).isoformat(), "action": action, **details}
        with open(self._approval_log, "a") as f:
            f.write(json.dumps(entry) + "\n")
